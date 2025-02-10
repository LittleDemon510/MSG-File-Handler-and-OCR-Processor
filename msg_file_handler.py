import os
import shutil
import re
import time
import datetime
from enum import Enum
from extract_msg import Message
from openpyxl import Workbook
from openpyxl import load_workbook
import uuid
import base64
import requests
import logging
 
 
 
# Traverses a folder and all its subfolders
def traverse(path):
    logger.debug('Count of files for folder ' + path + ' is ' + str(len(os.listdir(path))))
    for entry in os.scandir(path):
 
        if entry.is_dir(follow_symlinks=False):
            yield from traverse(entry.path)
        else:
            yield entry
 
# Loop through each msg file, check for a txt counterpart.
# If a txt counterpart is present append data of the txt file to the excel
# Else apply the necessary blockers with prefixes and profixes, cache the content as txt file, and then append to excel
def loop_msg(path,ws):
    count = 0
    parent_folder = path
    # Store duplicates of invoice and object numbers in a separate worksheet within the same Excel workbook.
    new_sheet = wb.create_sheet('Duplicates')
    new_sheet.append({'A': 'Content', 'B':'Results','C':'Duplicate Multiple Invoice numbers', 'D':'Duplicate Object numbers'})
    for entry in traverse(path):
        if entry.name.endswith('.msg'):
            new_path = rename(entry)
            if new_path:
                entry = new_path
            dirname = os.path.dirname(entry.replace(path + '\\',''))
                # Check if 'entry' is an instance of os.DirEntry (which represents a directory entry)
                # If true, assign 'entry.path' (the path to the directory entry) to 'msg_path'
                # If false, assign 'entry' (assumed to be a path string) directly to 'msg_path'
            msg_path = entry.path if isinstance(entry, os.DirEntry) else entry
            logger.debug(f'Message path is {msg_path}')
            logger.debug(f'parent_folder is {parent_folder}')
                # Open the file at 'msg_path' in binary read mode ('rb')
                # Use a 'with' statement to ensure the file is properly closed after reading
            with open(msg_path, 'rb') as f:
                msg = Message(f)
                subject = msg.subject
                body = msg.body
                subject_and_body = f"Subject: {subject}\n{body}"
                invoice_number = wrap_pattern(subject_and_body, pattern_to_wrap =r'\b2100\d{6}\b', prefix = "<Start:invoice_number>", postfix = "<End>")
                subject_and_body = invoice_number['replaced_text']
                duplicate_invoice_number = str(invoice_number['duplicates'])
                object_number = wrap_pattern(subject_and_body, pattern_to_wrap= r'M-\d{6}', prefix = "<Start:object_number>", postfix = "<End>")
                object_number = wrap_pattern(object_number['replaced_text'], pattern_to_wrap= r'\d{3}-\d{7}/\d{2}', prefix = "<Start:object_number>", postfix = "<End>")
                subject_and_body = object_number['replaced_text']
                duplicates_object_number = str(object_number['duplicates'])
                subject_and_body = replace_Disclaimer(subject_and_body)  
                subject_and_body += str(attachment(msg, msg_path, parent_folder))
            new_sheet.append([subject_and_body, duplicate_invoice_number, duplicates_object_number])

#Rename files with filenames that are longer than 100 characters        
def rename(entry):
    file_extension = ""
    if entry.name.endswith('.msg'): file_extension = '.msg'
    else: file_extension = '.txt'
    threshold = 100
    limit = 64
    entry_name = os.path.basename(entry).split('.')[0]
    logger.debug(f'Entry Name is {entry_name}')
    if len(entry.name) > threshold:
        logger.debug(f'Entry name is longer than {threshold} characters: {entry.path}')
        try:
            # Generate new filename with a UUID
            new_filename = entry.name[:limit] + str(uuid.uuid4()) + file_extension
            new_file_path = os.path.join(os.path.dirname(entry.path), new_filename)
 
            # Use '\\?\' prefix to handle long paths in Windows
            original_path_with_prefix = r"\\?\{}".format(os.path.abspath(entry.path))
            new_file_path_with_prefix = r"\\?\{}".format(os.path.abspath(new_file_path))
 
            # Check if the directory of the new file path exists
            if not os.path.exists(os.path.dirname(new_file_path_with_prefix)):
                logger.debug(f"Directory does not exist: {os.path.dirname(new_file_path_with_prefix)}")
            else:
                # Move the file to the new path
                shutil.move(original_path_with_prefix, new_file_path_with_prefix)
                logger.debug(f'Successfully renamed {original_path_with_prefix} to {new_file_path_with_prefix}')
                return new_file_path  # Return the new path after renaming
        except Exception as e:
            logger.error(f'Error renaming {entry.path}: {e}')
            return None
    else:
        return entry.path
 
# When an Invoice or object number is found, add prefix and postfix and returns the text as well as displays the number of duplicates
def wrap_pattern(text,pattern_to_wrap, prefix, postfix):
    pattern_counter = {}
    result = {
        'replaced_text': '',
        'duplicates': ['None']
    }
    matched_pattern =  re.findall(pattern_to_wrap, text)
    logger.debug(f'Matched pattern {pattern_to_wrap} is {matched_pattern}')
    # Check for duplicates of invoice numbers and object numbers.
    for match in matched_pattern:
        logger.debug(f'Match is {match}')
        key = match
        if key in pattern_counter.keys():
            pattern_counter.update({key:pattern_counter[key]+1})
        else:
            pattern_counter.update({key: 1})
    if len(pattern_counter.keys())>2:
        duplicate = list(pattern_counter.items())
        if len(duplicate) != 0:
            result['duplicates'] = duplicate
 
    result["replaced_text"] = wrapping(text,pattern_to_wrap, prefix, postfix)
    return result
 
def wrapping(text,pattern_to_wrap, prefix, postfix):
    regex = re.compile(pattern_to_wrap, re.IGNORECASE)
    return regex.sub(lambda match: prefix + match.group(0) + postfix, text)
 
# Checks for the pattern and when a pattern is found adds prefix and postfix and returns the text
def replace_Disclaimer(text):
    patterns = [r'External Email: Be cautious about the sender email address, attachments and links\. If uncertain use Report Message button\.',
    r'This is an external email\. Do you know who has sent it\? Can you be sure that any links and attachments contained within it are safe\? If in any doubt, use the Report Message button in your Outlook client to report this mail\.',
    r'This is an external email\.Do you know who has sent it\? Can you be sure that any links and attachments contained within it are safe\? If in any doubt, use the “Report Message” button in your Outlook client to report this mail\.',
    r'ACHTUNG: Diese E-Mail stammt von einem externen Kontakt\. Bitte gehen Sie mit Anhängen oder enthaltenen Links vorsichtig um\.',
    r'CYBER SECURITY WARNING: This email is from an external source - be careful of attachments and links\. Please follow the Cyber Code and report suspicious emails\.']
 
    count = 0
    pattern_match_disclaimer = ''
    for pattern in patterns:
        match = re.search(pattern,text)
        if(match):
            pattern_match_disclaimer = wrapping(text, pattern, prefix = "<Start:Disclaimer>", postfix = "<End>")
            count += 1
    logger.debug(f'Number of times patterns were matched and replaced {count}')
    return pattern_match_disclaimer
 
# Check for attachments witnin the msg file, if attachment is a pdf, encode it to Base64 and sent it to OCR Textract
# Get the OCR response, unpack it, and append the text from OCR to the content for the excel.
def attachment(msg, path, parent_folder):
    if(msg.attachments):
        for attachment in msg.attachments:
            file_name =  os.path.basename(path)
            logger.debug(f'Attachment type is : {type(attachment)} location is {file_name}')
            change_extension = file_name.split('.')[0] + ".txt"
            check_file_name = os.path.join(parent_folder, change_extension)
            logger.debug(f'File name to be checked: {check_file_name}')
            if os.path.isfile(check_file_name):
                logger.debug("OCR has checked these files before")
                with open(check_file_name,'r', encoding = 'utf-8') as f:
                    results = f.read()
                    logger.debug(f'Content from existing txt file is {results}')
                    return results
            else:
                attachment_name = attachment.longFilename
                if attachment_name is not None:
                    file_extension = attachment_name.split('.')[-1]
                    logger.debug(f'Attachment long file name is {attachment_name} and extension is {file_extension}')
                    data = attachment.data #Binary data
                    if attachment_name.lower().endswith(('pdf','jpg','png')):
                        logger.debug(f'File {attachment.longFilename} is an accepted file')
                        base64_string = base64.b64encode(data).decode('utf-8')
                        url = 'dummy url'
                        proxies={
                            'dummy proxy',
                            'dummy proxy'}
                        payload = {
                            "name": file_name,
                            "extension": file_extension,
                            "content": base64_string
                            }
                        headers = {
                            'Content-Type': 'application/json'
                            }
                        try:
                            response = requests.post(url, json = payload, headers=headers, proxies = proxies)
                            logger.debug(f'response is {response}')
                            results = response.json()
                            text = results['pages']
                            text_data = [td.get('text', None)for td in text]
                            logger.debug(f'Text from OCR is {text_data}')
                            return text_data
                        except Exception as e:
                            logger.error(f'An unexpected error occured: {e}')
                    else:
                        logger.error("File is of an unacceptable filetype")
 
######################### -= MAIN =- #########################
#To calculate total run time of the script
startTime = time.time()
current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
logging.basicConfig(filename=f'I_{current_time}.log', format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO,   datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger("I")
# Path to the root folder containing all the emails
# Ask for the file location
msg_folder_path = input(r"Enter the path to MASTER file: ")
logger.info(f"Selected file path is:{msg_folder_path}")
# Save a copy of the template file with the desired name
excel_file_name = input('Name for the Excel file (empty to autogenerate): ')
if excel_file_name.strip() == '':
    excel_file_name = 'trainingDataSet_' + current_time
excel_file_name = excel_file_name + '.xlsx'
shutil.copy('training_template.xlsx',excel_file_name)
 
# Load the emails into excel sheet
wb = load_workbook(excel_file_name)
ws = wb['DATA']
loop_msg(msg_folder_path,ws)
 
 
# Refresh the Pivot table
ws = wb['PIVOT']
pivot = ws._pivots[0]
pivot.cache.refreshOnLoad = True
 
# Save the file
wb.save(excel_file_name)
wb.close()
 
#Calculate total run time of the script
endTime = time.time()
elapsedTime = endTime - startTime
logger.info('DONE! Total elpased time in seconds: %s', elapsedTime)import os
import shutil
import re
import time
import datetime
from enum import Enum
from extract_msg import Message
from openpyxl import Workbook
from openpyxl import load_workbook
import uuid
import base64
import requests
import logging
 
 
 
# Traverses a folder and all its subfolders
def traverse(path):
    logger.debug('Count of files for folder ' + path + ' is ' + str(len(os.listdir(path))))
    for entry in os.scandir(path):
 
        if entry.is_dir(follow_symlinks=False):
            yield from traverse(entry.path)
        else:
            yield entry
 
# Loop through each msg file, check for a txt counterpart.
# If a txt counterpart is present append data of the txt file to the excel
# Else apply the necessary blockers with prefixes and profixes, cache the content as txt file, and then append to excel
def loop_msg(path,ws):
    count = 0
    parent_folder = path
    # Store duplicates of invoice and object numbers in a separate worksheet within the same Excel workbook.
    new_sheet = wb.create_sheet('Duplicates')
    new_sheet.append({'A': 'Content', 'B':'Results','C':'Duplicate Multiple Invoice numbers', 'D':'Duplicate Object numbers'})
    for entry in traverse(path):
        if entry.name.endswith('.msg'):
            new_path = rename(entry)
            if new_path:
                entry = new_path
            dirname = os.path.dirname(entry.replace(path + '\\',''))
                # Check if 'entry' is an instance of os.DirEntry (which represents a directory entry)
                # If true, assign 'entry.path' (the path to the directory entry) to 'msg_path'
                # If false, assign 'entry' (assumed to be a path string) directly to 'msg_path'
            msg_path = entry.path if isinstance(entry, os.DirEntry) else entry
            logger.debug(f'Message path is {msg_path}')
            logger.debug(f'parent_folder is {parent_folder}')
                # Open the file at 'msg_path' in binary read mode ('rb')
                # Use a 'with' statement to ensure the file is properly closed after reading
            with open(msg_path, 'rb') as f:
                msg = Message(f)
                subject = msg.subject
                body = msg.body
                subject_and_body = f"Subject: {subject}\n{body}"
                invoice_number = wrap_pattern(subject_and_body, pattern_to_wrap =r'\b2100\d{6}\b', prefix = "<Start:invoice_number>", postfix = "<End>")
                subject_and_body = invoice_number['replaced_text']
                duplicate_invoice_number = str(invoice_number['duplicates'])
                object_number = wrap_pattern(subject_and_body, pattern_to_wrap= r'M-\d{6}', prefix = "<Start:object_number>", postfix = "<End>")
                object_number = wrap_pattern(object_number['replaced_text'], pattern_to_wrap= r'\d{3}-\d{7}/\d{2}', prefix = "<Start:object_number>", postfix = "<End>")
                subject_and_body = object_number['replaced_text']
                duplicates_object_number = str(object_number['duplicates'])
                subject_and_body = replace_Disclaimer(subject_and_body)  
                subject_and_body += str(attachment(msg, msg_path, parent_folder))
            new_sheet.append([subject_and_body, duplicate_invoice_number, duplicates_object_number])

#Rename files with filenames that are longer than 100 characters        
def rename(entry):
    file_extension = ""
    if entry.name.endswith('.msg'): file_extension = '.msg'
    else: file_extension = '.txt'
    threshold = 100
    limit = 64
    entry_name = os.path.basename(entry).split('.')[0]
    logger.debug(f'Entry Name is {entry_name}')
    if len(entry.name) > threshold:
        logger.debug(f'Entry name is longer than {threshold} characters: {entry.path}')
        try:
            # Generate new filename with a UUID
            new_filename = entry.name[:limit] + str(uuid.uuid4()) + file_extension
            new_file_path = os.path.join(os.path.dirname(entry.path), new_filename)
 
            # Use '\\?\' prefix to handle long paths in Windows
            original_path_with_prefix = r"\\?\{}".format(os.path.abspath(entry.path))
            new_file_path_with_prefix = r"\\?\{}".format(os.path.abspath(new_file_path))
 
            # Check if the directory of the new file path exists
            if not os.path.exists(os.path.dirname(new_file_path_with_prefix)):
                logger.debug(f"Directory does not exist: {os.path.dirname(new_file_path_with_prefix)}")
            else:
                # Move the file to the new path
                shutil.move(original_path_with_prefix, new_file_path_with_prefix)
                logger.debug(f'Successfully renamed {original_path_with_prefix} to {new_file_path_with_prefix}')
                return new_file_path  # Return the new path after renaming
        except Exception as e:
            logger.error(f'Error renaming {entry.path}: {e}')
            return None
    else:
        return entry.path
 
# When an Invoice or object number is found, add prefix and postfix and returns the text as well as displays the number of duplicates
def wrap_pattern(text,pattern_to_wrap, prefix, postfix):
    pattern_counter = {}
    result = {
        'replaced_text': '',
        'duplicates': ['None']
    }
    matched_pattern =  re.findall(pattern_to_wrap, text)
    logger.debug(f'Matched pattern {pattern_to_wrap} is {matched_pattern}')
    # Check for duplicates of invoice numbers and object numbers.
    for match in matched_pattern:
        logger.debug(f'Match is {match}')
        key = match
        if key in pattern_counter.keys():
            pattern_counter.update({key:pattern_counter[key]+1})
        else:
            pattern_counter.update({key: 1})
    if len(pattern_counter.keys())>2:
        duplicate = list(pattern_counter.items())
        if len(duplicate) != 0:
            result['duplicates'] = duplicate
 
    result["replaced_text"] = wrapping(text,pattern_to_wrap, prefix, postfix)
    return result
 
def wrapping(text,pattern_to_wrap, prefix, postfix):
    regex = re.compile(pattern_to_wrap, re.IGNORECASE)
    return regex.sub(lambda match: prefix + match.group(0) + postfix, text)
 
# Checks for the pattern and when a pattern is found adds prefix and postfix and returns the text
def replace_Disclaimer(text):
    patterns = [r'External Email: Be cautious about the sender email address, attachments and links\. If uncertain use Report Message button\.',
    r'This is an external email\. Do you know who has sent it\? Can you be sure that any links and attachments contained within it are safe\? If in any doubt, use the Report Message button in your Outlook client to report this mail\.',
    r'This is an external email\.Do you know who has sent it\? Can you be sure that any links and attachments contained within it are safe\? If in any doubt, use the “Report Message” button in your Outlook client to report this mail\.',
    r'ACHTUNG: Diese E-Mail stammt von einem externen Kontakt\. Bitte gehen Sie mit Anhängen oder enthaltenen Links vorsichtig um\.',
    r'CYBER SECURITY WARNING: This email is from an external source - be careful of attachments and links\. Please follow the Cyber Code and report suspicious emails\.']
 
    count = 0
    pattern_match_disclaimer = ''
    for pattern in patterns:
        match = re.search(pattern,text)
        if(match):
            pattern_match_disclaimer = wrapping(text, pattern, prefix = "<Start:Disclaimer>", postfix = "<End>")
            count += 1
    logger.debug(f'Number of times patterns were matched and replaced {count}')
    return pattern_match_disclaimer
 
# Check for attachments witnin the msg file, if attachment is a pdf, encode it to Base64 and sent it to OCR Textract
# Get the OCR response, unpack it, and append the text from OCR to the content for the excel.
def attachment(msg, path, parent_folder):
    if(msg.attachments):
        for attachment in msg.attachments:
            file_name =  os.path.basename(path)
            logger.debug(f'Attachment type is : {type(attachment)} location is {file_name}')
            change_extension = file_name.split('.')[0] + ".txt"
            check_file_name = os.path.join(parent_folder, change_extension)
            logger.debug(f'File name to be checked: {check_file_name}')
            if os.path.isfile(check_file_name):
                logger.debug("OCR has checked these files before")
                with open(check_file_name,'r', encoding = 'utf-8') as f:
                    results = f.read()
                    logger.debug(f'Content from existing txt file is {results}')
                    return results
            else:
                attachment_name = attachment.longFilename
                if attachment_name is not None:
                    file_extension = attachment_name.split('.')[-1]
                    logger.debug(f'Attachment long file name is {attachment_name} and extension is {file_extension}')
                    data = attachment.data #Binary data
                    if attachment_name.lower().endswith(('pdf','jpg','png')):
                        logger.debug(f'File {attachment.longFilename} is an accepted file')
                        base64_string = base64.b64encode(data).decode('utf-8')
                        url = 'dummy url'
                        proxies={
                            'dummy proxy',
                            'dummy proxy'}
                        payload = {
                            "name": file_name,
                            "extension": file_extension,
                            "content": base64_string
                            }
                        headers = {
                            'Content-Type': 'application/json'
                            }
                        try:
                            response = requests.post(url, json = payload, headers=headers, proxies = proxies)
                            logger.debug(f'response is {response}')
                            results = response.json()
                            text = results['pages']
                            text_data = [td.get('text', None)for td in text]
                            logger.debug(f'Text from OCR is {text_data}')
                            return text_data
                        except Exception as e:
                            logger.error(f'An unexpected error occured: {e}')
                    else:
                        logger.error("File is of an unacceptable filetype")
 
######################### -= MAIN =- #########################
#To calculate total run time of the script
startTime = time.time()
current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
logging.basicConfig(filename=f'I_{current_time}.log', format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO,   datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger("I")
# Path to the root folder containing all the emails
# Ask for the file location
msg_folder_path = input(r"Enter the path to MASTER file: ")
logger.info(f"Selected file path is:{msg_folder_path}")
# Save a copy of the template file with the desired name
excel_file_name = input('Name for the Excel file (empty to autogenerate): ')
if excel_file_name.strip() == '':
    excel_file_name = 'trainingDataSet_' + current_time
excel_file_name = excel_file_name + '.xlsx'
shutil.copy('training_template.xlsx',excel_file_name)
 
# Load the emails into excel sheet
wb = load_workbook(excel_file_name)
ws = wb['DATA']
loop_msg(msg_folder_path,ws)
 
 
# Refresh the Pivot table
ws = wb['PIVOT']
pivot = ws._pivots[0]
pivot.cache.refreshOnLoad = True
 
# Save the file
wb.save(excel_file_name)
wb.close()
 
#Calculate total run time of the script
endTime = time.time()
elapsedTime = endTime - startTime
logger.info('DONE! Total elpased time in seconds: %s', elapsedTime)
